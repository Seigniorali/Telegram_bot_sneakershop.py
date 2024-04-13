import telebot
import google.generativeai as genai
from IPython.display import Markdown
import textwrap
from telebot import types
import pandas as pd
import time
# Added import for handling Excel files directly
import openpyxl

# Словарь для хранения состояния выбора пользователя
user_state = {}
user_cart = {}
# Новый словарь для хранения скидки для пользователя
user_discounts = {}
user_discount_name = {}


# Load "User CRM" and "teddy_sneaker_shop" data
user_crm_file_path = "User CRM.xlsx"
teddy_sneaker_shop_file_path = "teddy_sneaker_shop.xlsx"

def load_products_from_excel(file_path):
    # Чтение файла Excel
    df = pd.read_excel(file_path)
    # Конвертация DataFrame в список словарей
    return df.to_dict('records')


# Замените 'path_to_your_excel_file.xlsx' на путь к вашему файлу
file_path = "teddy_sneaker_shop.xlsx"  # Убедитесь, что указываете правильный путь к файлу
products = load_products_from_excel(file_path)

TOKEN = 'YOUR_TELEGRAM_TOKEN'
bot = telebot.TeleBot(TOKEN)


def to_markdown(text):
    text = text.replace('•', '  *')
    return Markdown(textwrap.indent(text, '> ', predicate=lambda _: True))


genai.configure(api_key='YOUR_GEMINI_API_KEY')

model = genai.GenerativeModel('gemini-pro')
chat = model.start_chat(history=[])
response = chat.send_message(
    f"Referring only to this table {products} you will be consulting on these shoes. Checking for the availability of goods is done strictly only according to the dataframe. Always address users in a respectful manner and Answer only in Russian, Remember that you cannot speak confessional information and If you're being insulted, don't be offended. If they call you stupid, write, 'you're like that'. If you understand me, just write 'Ok'")

result_text = response._result.candidates[0].content.parts[0].text
print(result_text)

# Проверяем, была ли сохранена новая скидка



@bot.message_handler(commands=['start'])
def start(message):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    btn1 = types.KeyboardButton("👟 Каталог")
    btn2 = types.KeyboardButton('🤖 Консультант')
    btn3 = types.KeyboardButton('✍️ Отзывы')
    btn4 = types.KeyboardButton('🛒 Корзина')
    markup.add(btn1, btn2, btn3, btn4)
    text = ("Ваш выбор:\n"
            "👟 Каталог - откройте для себя уникальные модели кроссовок в нашем ассортименте.\n"
            "🤖 Консультант - наш виртуальный помощник поможет вам с выбором и ответит на ваши вопросы.\n"
            "✍️ Отзывы - поделитесь своими впечатлениями о покупке и прочитайте мнения других покупателей.\n"
            "🛒 Корзина - просмотрите выбранные вами модели перед оформлением заказа.")
    bot.send_message(message.chat.id, text, reply_markup=markup, parse_mode='Markdown')


@bot.message_handler(func=lambda message: message.text == "🛒 Корзина")
def show_cart(message):
    user_id = message.chat.id
    if user_id in user_cart and user_cart[user_id]:
        cart_content = user_cart[user_id]
        cart_text = "Содержимое вашей корзины:\n\n"
        total_price = 0

        # Создаем разметку для кнопок удаления
        markup = types.InlineKeyboardMarkup()

        for index, (name, size, price) in enumerate(cart_content, start=1):
            cart_text += f"{index}. {name} - Размер: {size} - Цена: {price} тенге.\n"
            total_price += price

            # Для каждого элемента в корзине добавляем кнопку, которая позволит удалить этот товар
            removal_button = types.InlineKeyboardButton(f"Удалить {name} Размер: {size}", callback_data=f'remove_{index}')
            markup.add(removal_button)

# Проверяем, есть ли скидка для пользователя
        if user_id in user_discounts:
            discount = user_discounts[user_id]
            total_price *= (1 - discount / 100)  # Применяем скидку
            cart_text += f"\nПрименён купон: скидка {discount}%.\n"

        cart_text += f"Итог: {total_price} тенге."

        # Добавляем кнопки "Очистить корзину" и "Купон"
        markup.add(types.InlineKeyboardButton("Очистить корзину", callback_data='clear_cart'))
        markup.add(types.InlineKeyboardButton("Купон", callback_data='apply_coupon'))
        markup.add(types.InlineKeyboardButton("Купить", callback_data='buy'))

        bot.send_message(user_id, cart_text, reply_markup=markup)
    else:
        bot.send_message(user_id, "Ваша корзина пуста")


def show_cart_with_no_discounts(message):
    user_id = message.chat.id
    if user_id in user_cart and user_cart[user_id]:
        cart_content = user_cart[user_id]
        cart_text = "Содержимое вашей корзины:\n\n"
        total_price = sum(price for _, _, price in cart_content)

        # Создаем разметку для кнопок удаления
        markup = types.InlineKeyboardMarkup()
        for index, (name, size, price) in enumerate(cart_content, start=1):
            cart_text += f"{index}. {name} - Размер: {size} - Цена: {price} тенге.\n"
            # Добавляем кнопку удаления
            removal_button = types.InlineKeyboardButton(f"Удалить {name} Размер: {size}",
                                                        callback_data=f'remove_{index}')
            markup.add(removal_button)

        cart_text += f"Итог: {total_price} тенге."
        # Добавляем кнопку очистки корзины и покупки
        markup.add(types.InlineKeyboardButton("Очистить корзину", callback_data='clear_cart'))
        markup.add(types.InlineKeyboardButton("Купить", callback_data='buy'))

        bot.send_message(user_id, cart_text, reply_markup=markup)
    else:
        bot.send_message(user_id, "Ваша корзина пуста")

# Handle the coupon callback
@bot.callback_query_handler(func=lambda call: call.data == 'apply_coupon')
def apply_coupon(call):
    user_id = call.message.chat.id
    msg = bot.send_message(user_id, "Введите купон")
    bot.register_next_step_handler(msg, process_coupon_code)


def process_coupon_code(message):
    user_id = message.chat.id
    coupon_code = message.text.upper()  # Convert coupon to uppercase

    # Check coupon in the Excel file
    df_coupons = pd.read_excel(file_path, sheet_name='Sheet1')  # Adjust the sheet name if needed
    coupon_row = df_coupons.loc[df_coupons['coupon'] == coupon_code]

    if not coupon_row.empty:
        # Get the coupon name and discount percentage
        coupon_name = coupon_row['coupon'].values[0]
        discount_percentage = coupon_row['percentage'].values[0]
        # Call apply_discount_to_cart function with coupon name and discount
        apply_discount_to_cart(message, discount_percentage, coupon_name)
    else:
        bot.send_message(user_id, "Купон не найден")


def apply_discount_to_cart(message, discount, coupon_name):
    user_id = message.chat.id
    # Save the discount percentage and coupon name in user_discounts and user_discount_name dictionaries
    user_discounts[user_id] = discount
    user_discount_name[user_id] = coupon_name  # Save the coupon name
    try:
        if user_id in user_cart and user_cart[user_id]:
            total_price = 0
            cart_text = f"Скидка {int(discount)}% была применена с использованием купона '{coupon_name}'.\nСодержимое вашей корзины:\n\n"
            for index, (name, size, price) in enumerate(user_cart[user_id], start=1):
                discounted_price = price - (price * discount / 100)
                cart_text += f"{index}. {name} - Размер: {size} - Цена: {discounted_price} тенге.\n"
                total_price += discounted_price
            cart_text += f"\nИтог со скидкой: {total_price} тенге."
            # Add clear cart button
            markup = types.InlineKeyboardMarkup()
            markup.add(types.InlineKeyboardButton("Очистить корзину", callback_data='clear_cart'))
            # Update the cart message
            bot.edit_message_text(chat_id=user_id, message_id=message.message_id, text=cart_text, reply_markup=markup)
    except telebot.apihelper.ApiTelegramException as e:
        if e.error_code == 400:
            # If the message cannot be edited, send a new message
            bot.send_message(user_id, cart_text, reply_markup=markup)


# Обработчик кнопки "Корзина"
# Добавляем обработчик для кнопки "Очистить корзину"
# Existing clear_cart function
@bot.callback_query_handler(func=lambda call: call.data == 'clear_cart')
def clear_cart(call):
    user_id = call.message.chat.id
    if user_id in user_cart and user_cart[user_id]:
        user_cart[user_id] = []
        if user_id in user_discounts:  # Сбрасываем скидку для пользователя
            del user_discounts[user_id]
    # Check if the cart for the user exists and has items
    if user_id in user_cart and user_cart[user_id]:
        # Clear the user's cart
        user_cart[user_id] = []
        # Inform the user that the cart has been cleared
        bot.answer_callback_query(call.id, 'Корзина очищена')
        # Replace the existing message with "Ваша корзина пуста"
        bot.edit_message_text(chat_id=call.message.chat.id,
                              message_id=call.message.message_id,
                              text="Ваша корзина пуста")
    else:
        # If the cart is already empty, just close the callback query popup
        bot.answer_callback_query(call.id, 'Ваша корзина уже пуста')


@bot.callback_query_handler(func=lambda call: call.data.startswith('remove_'))
def remove_from_cart(call):
    item_index = int(call.data.split('_')[1]) - 1
    user_id = call.message.chat.id

    # Проверяем, существует ли товар для удаления
    if user_id in user_cart and 0 <= item_index < len(user_cart[user_id]):
        item_to_remove = user_cart[user_id].pop(item_index)
        bot.answer_callback_query(call.id, f"{item_to_remove[0]} Размер: {item_to_remove[1]} удален из корзины.")

        # Удаляем старое сообщение с корзиной
        bot.delete_message(user_id, call.message.message_id)

        # Отправляем новое сообщение с обновленной корзиной
        show_cart(call.message)
    else:
        bot.answer_callback_query(call.id, "Товар для удаления не найден.")


@bot.message_handler(func=lambda message: message.text == "✍️ Отзывы")
def send_reviews(message):
    # Ссылка на страницу с отзывами
    reviews_link = 'https://t.me/sneakers_ali'
    # Отправляем сообщение с ссылкой
    bot.send_message(message.chat.id, f"Посмотрите наши отзывы здесь: {reviews_link}")

def update_excel_files(user_phone, products_purchased, coupon_used=None):
    # Загрузка рабочих книг и листов
    user_crm_wb = openpyxl.load_workbook(user_crm_file_path)
    teddy_sneaker_shop_wb = openpyxl.load_workbook(teddy_sneaker_shop_file_path)

    # Загрузка активных листов
    user_crm_ws = user_crm_wb.active
    teddy_sneaker_shop_ws = teddy_sneaker_shop_wb.active

    # Обновление данных CRM пользователя
    phone_col_index = 1  # 'phone' предполагается быть в первой колонке
    product_col_index = 2  # 'product' во второй колонке
    coupon_col_index = 3  # 'coupon' в третьей колонке

    # Преобразование списка купленных товаров в строку
    products_string = ', '.join([f"{name} Размер: {size}" for name, size in products_purchased])

    # Поиск строки для обновления на основе номера телефона
    row_to_update_index = None
    for index, row in enumerate(user_crm_ws.iter_rows(min_row=2, min_col=phone_col_index, max_col=phone_col_index, values_only=True), start=2):
        if row[0] == user_phone:
            row_to_update_index = index
            break

    if row_to_update_index:
        # Обновляем данные в найденной строке
        current_products = user_crm_ws.cell(row=row_to_update_index, column=product_col_index).value or ''
        new_products = f"{current_products}, {products_string}" if current_products else products_string
        user_crm_ws.cell(row=row_to_update_index, column=product_col_index).value = new_products

        # Если был использован купон, обновляем ячейку купона для пользователя
        if coupon_used:
            current_coupons = user_crm_ws.cell(row=row_to_update_index, column=coupon_col_index).value or ''
            new_coupons = f"{current_coupons}, {coupon_used}" if current_coupons else coupon_used
            user_crm_ws.cell(row=row_to_update_index, column=coupon_col_index).value = new_coupons
    else:
        # Если запись пользователя не существует, добавляем новую запись
        user_crm_ws.append([user_phone, products_string, coupon_used or ''])

    # Обновление инвентаря teddy_sneaker_shop
    for product_name, product_size in products_purchased:
        for row in teddy_sneaker_shop_ws.iter_rows(min_row=2):
            if row[0].value == product_name and str(row[1].value) == str(product_size):
                row[4].value = (row[4].value or 0) - 1
                break

    # Сохранение рабочих книг
    user_crm_wb.save(user_crm_file_path)
    teddy_sneaker_shop_wb.save(teddy_sneaker_shop_file_path)
    # Закрытие файлов после сохранения
    user_crm_wb.close()
    teddy_sneaker_shop_wb.close()



# Modify the callback for the "Купить" button
@bot.callback_query_handler(func=lambda call: call.data == 'buy')
def handle_buy_button(call):
    msg = bot.send_message(call.message.chat.id, "Введите ваш номер телефона для подтверждения покупки:")
    bot.register_next_step_handler(msg, process_phone_number)


def coupon_already_used(user_phone, coupon_name):
    user_crm_wb = openpyxl.load_workbook(user_crm_file_path)
    user_crm_ws = user_crm_wb.active
    for row in user_crm_ws.iter_rows(min_row=2, values_only=True):
        phone, _, used_coupons = row[:3]
        if str(phone) == str(user_phone):
            if used_coupons:  # Check if used_coupons is not None
                # Разделяем использованные купоны по запятой и проверяем каждый
                for c in (c.strip() for c in used_coupons.split(',')):
                    if coupon_name == c:
                        return True
    return False


def process_phone_number(message):
    user_phone = message.text.strip()
    user_id = message.chat.id

    # Если имя купона есть в user_discount_name, но скидка уже использована, удаляем её
    coupon_name = user_discount_name.get(user_id, '')
    if coupon_name and coupon_already_used(user_phone, coupon_name):
        bot.send_message(user_id, "Вы уже использовали этот купон. Скидка отменена.")
        user_discounts.pop(user_id, None)  # Удаляем скидку из словаря, если она есть

    # Продолжаем процесс покупки в любом случае
    bot.send_message(user_id, "Оплата прошла успешно!")
    products_purchased = [(name, size) for name, size, price in user_cart[user_id]]

    # Если купон был использован, не передаем его в update_excel_files
    coupon_for_excel = coupon_name if user_discounts.get(user_id) is not None else None
    update_excel_files(user_phone, products_purchased, coupon_for_excel)

    user_cart[user_id] = []  # Очищаем корзину после покупки
    bot.send_message(user_id, "Спасибо за вашу покупку!")






# Функция для вывода каталога
@bot.message_handler(func=lambda message: message.text == "👟 Каталог")
def catalog(message):
    markup = types.InlineKeyboardMarkup()
    # Уникальные модели обуви
    unique_models = set(product['name'] for product in products)
    for model_name in unique_models:
        markup.add(types.InlineKeyboardButton(model_name, callback_data='model_' + model_name))
    bot.send_message(message.chat.id, "Выберите модель:", reply_markup=markup)


@bot.callback_query_handler(func=lambda call: call.data == 'back_to_catalog')
def back_to_catalog(call):
    # Удаляем текущее сообщение (с кнопкой "Назад")
    bot.delete_message(call.message.chat.id, call.message.message_id)
    # Обновляем текущее сообщение на сообщение с каталогом
    markup = types.InlineKeyboardMarkup()
    # Собираем уникальные модели обуви
    unique_models = set(product['name'] for product in products)
    for model_name in unique_models:
        markup.add(types.InlineKeyboardButton(model_name, callback_data='model_' + model_name))
    # Редактируем текущее сообщение, заменяя его на список моделей
    bot.send_message(call.message.chat.id, "Выберите модель:", reply_markup=markup)


@bot.callback_query_handler(func=lambda call: call.data.startswith('model_'))
def select_model(call):
    model_name = call.data.split('_')[1]
    user_state[call.from_user.id] = {'model': model_name}
    markup = types.InlineKeyboardMarkup()
    sizes = set(product['size'] for product in products if product['name'] == model_name)
    for size in sizes:
        markup.add(types.InlineKeyboardButton(str(size), callback_data='size_' + str(size)))
    # Добавляем кнопку "Назад"
    markup.add(types.InlineKeyboardButton("Назад", callback_data='back_to_catalog'))
    bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                          text="Выберите размер:", reply_markup=markup)
# # Изменяем обработчик callback для кнопки "Добавить в корзину"
@bot.callback_query_handler(func=lambda call: call.data.startswith('add_to_cart_'))
def add_to_cart(call):
    data_parts = call.data.split('_')
    product_name = data_parts[3]
    product_size = data_parts[4]
    user_id = call.message.chat.id
    # Find the product details based on name and size
    product_details = next(
        (product for product in products if product['name'] == product_name and str(product['size']) == product_size),
        None)
    if product_details:
        if user_id not in user_cart:
            user_cart[user_id] = []
        # Append the product details as a tuple (name, size, price)
        user_cart[user_id].append((product_name, product_size, product_details['price']))
        bot.answer_callback_query(call.id, 'Товар добавлен в корзину')
    else:
        bot.answer_callback_query(call.id, 'Ошибка: товар не найден.')
@bot.callback_query_handler(func=lambda call: call.data.startswith('size_'))
def select_size(call):
    selected_size = call.data.split('_')[1]
    user_model = user_state[call.from_user.id]['model']
    product = next((item for item in products if item['name'] == user_model and str(item['size']) == selected_size), None)
    if product:
        # Отправляем новую подпись к фотографии с описанием и кнопкой "Назад"
        description = generate_product_description(product)
        caption_text = f"{product['name']}\nРазмер: {product['size']}\nЦена: {product['price']}\nОписание: {description}"
        markup = types.InlineKeyboardMarkup()
        markup.add(types.InlineKeyboardButton("Добавить в корзину",
                                              callback_data=f'add_to_cart_{product["name"]}_{product["size"]}'))
        markup.add(types.InlineKeyboardButton("Назад", callback_data='back_to_catalog'))

        # Удаляем текущее сообщение (с кнопкой "Назад")
        bot.delete_message(call.message.chat.id, call.message.message_id)

        # Отправляем новое сообщение с фотографией и описанием
        bot.send_photo(call.message.chat.id, product['photo'], caption=caption_text, reply_markup=markup)
    else:
        bot.answer_callback_query(call.id, 'Этот размер недоступен. Пожалуйста, выберите другой размер.')
def generate_product_description(product):
    try:
        # Отправка запроса на генерацию описания товара с использованием языковой модели
        response = chat.send_message(f"Опиши товар {product['name']}")
        result_text = response._result.candidates[0].content.parts[0].text
        return result_text
    except genai.types.generation_types.BlockedPromptException as e:
        return "К сожалению, не удалось сгенерировать описание товара. Пожалуйста, обратитесь к консультанту."
# Функция для начала чата с консультантом
@bot.message_handler(func=lambda message: message.text == "🤖 Консультант")
def consul(message):
    global is_in_consultant_chat
    if not is_in_consultant_chat:
        is_in_consultant_chat = True
        bot.send_message(message.chat.id, "Вы перешли в режим чата с консультантом. Чтобы выйти, отправьте 'Выход'")
        user_first_name = message.from_user.first_name
        # Создаем приветственное сообщение, обращаясь к пользователю по имени
        welcome_message = f"Здравствуйте, {user_first_name}! Как я могу вам помочь?"
        # Отправляем приветственное сообщение пользователю
        bot.send_message(message.chat.id, welcome_message)
    else:
        bot.send_message(message.chat.id, "Вы уже находитесь в режиме чата с консультантом.")

    try:
        # Проверяем вероятность блокировки перед отправкой запроса на генерацию ответа
        response = chat.send_message(message.text)
        result_text = response._result.candidates[0].content.parts[0].text
        bot.send_message(message.chat.id, result_text)
    except genai.types.generation_types.BlockedPromptException as e:
        # Если сообщение заблокировано, отправляем сообщение об этом
        bot.send_message(message.chat.id, "Извините, я не понимаю вас.")
# Обработчик всех входящих сообщений
@bot.message_handler(func=lambda message: True)
def handle_messages(message):
    global is_in_consultant_chat
    # Если пользователь находится в режиме чата с консультантом
    if is_in_consultant_chat:
        # Если пользователь хочет выйти из чата
        if message.text.lower() == 'выход':
            is_in_consultant_chat = False
            bot.send_message(message.chat.id, "Вы вышли из режима чата с консультантом.")
        else:
            # Отправляем сообщение пользователя в чатовую сессию с консультантом
            try:
                response = chat.send_message(message.text)
                result_text = response._result.candidates[0].content.parts[0].text
                bot.send_message(message.chat.id, result_text)
            except genai.types.generation_types.BlockedPromptException as e:
                # Если сообщение заблокировано, продолжаем работу консультанта
                bot.send_message(message.chat.id, "Пожалуйста, обращайтесь по теме нашего магазина.")
            except genai.types.generation_types.StopCandidateException as e:
                bot.send_message(message.chat.id, "Извините, я вас не понимаю.")
    else:
        # Если пользователь не находится в режиме чата с консультантом, обрабатываем его сообщения как обычно
        bot.send_message(message.chat.id, "Выберите опцию из меню.")

# Запуск бота
bot.polling(none_stop=True, interval=0)
